import requests, \
    os, \
    json, \
    openpyxl as xl, \
    xlrd  as rd, \
    xlsxwriter, \
    csv , \
    re
from xlrd import open_workbook
from string import Template
from openpyxl import Workbook


def getOrgNames():
    with open("src/inputFile.txt") as fp:
        orgNames = fp.readlines()
        return orgNames

def extract_values(obj, key):
    """Pull all values of specified key from nested JSON."""
    arr = []

    def extract(obj, arr, key):
        """Recursively search for values of key in JSON tree."""
        if isinstance(obj, dict):
            #print("Key=",key)
            for k, v in obj.items():
                #print("k=",k,"v=",v)
               # if isinstance(v, (dict, list)):
                if isinstance(v, (dict)):
                    extract(v, arr, key)  # recurrsive search with sub item
                elif k == key:  # macthing with Key
                    arr.append(v)
        elif isinstance(obj, list):  # searching in the list
            for item in obj:
                extract(item, arr, key)
        return arr  # returing complete Array

    results = extract(obj, arr, key)  # obj is Data, arr is return result and key is "search keyword"
    return results

def extract_general_values(obj, key):
    """Pull all values of specified key from nested JSON."""
    arr = []

    def extract(obj, arr, key):
        """Recursively search for values of key in JSON tree."""
        if isinstance(obj, dict):
            #print("Key=",key)
            for k, v in obj.items():
                #print("k=",k,"v=",v)
                if isinstance(v, (dict, list)):
                    extract(v, arr, key)  # recurrsive search with sub item
                elif k == key:  # macthing with Key
                    arr.append(v)
        elif isinstance(obj, list):  # searching in the list
            for item in obj:
                extract(item, arr, key)
        return arr  # returing complete Array

    results = extract(obj, arr, key)  # obj is Data, arr is return result and key is "search keyword"
    return results

def runQuery(realQuery):
    request = requests.post('http://cf-api.apps-za.homedepot.com/graphql', json={'query': realQuery})
    if request.status_code == 200:
        return request.json()
    else:
        raise Exception("Query failed to run by returning code of {}. {}".format(request.status_code, realQuery))

def createNewFile(dirName, fileName, sheetName):
    os.chdir(dirName)
    # Call a Workbook() function of openpyxl
    # to create a new blank Workbook object
    wb = xl.Workbook()

    # Get workbook active sheet
    # from the active attribute
    sheet = wb.active

    # removing extra "//" from file name
    fileName = fileName.replace("//", "")

    # One can change the name of the title
    sheet.title = sheetName
    wb.save(fileName)
    #wb.close()

def getParameterDetails():
    cwd = os.getcwd()
    dirName = "parameter"
    fileName = "ExpSubExp.json"
    os.chdir(dirName)
    subExp = []
    exp = "Customer Order Management"

    # Loading Parameter details into Json Object
    with open(fileName) as jsonFile:
        jd = json.load(jsonFile)

    # Fetching all the Org name respectively
    orgArray = extract_values(jd,'orgName')

    # Fetching Sub Exp name from Json Object
    for key, value in jd.items():
        subExp.append(key)

    os.chdir(cwd)
    return exp, subExp , orgArray

def getNextAvailableCell():
    dirName = "outputFiles"
    fileName = "PCF_App_Details.xlsx"
    sheetName = "Application_Details"
    cwd = os.getcwd()
    os.chdir("../" + dirName)
    wb = rd.open_workbook(fileName)
    sheet = wb.sheet_by_index(0)
    # sheet.cell_value(0, 0)
    maxCols = 0
    totRows = sheet.nrows
    totCol = sheet.ncols
    try:
        for rw in range(totRows):
            for cl in range(totCol):
                # if rw == 0 and cl == 0:
                #    continue
                if sheet.cell_value(rw, cl) != "":
                    if cl >= maxCols:
                        maxCols = cl + 1
    except IndexError:
        pass
    os.chdir(cwd)
    return maxCols

def writingFile(exp,orgName, subExp, managerNames, appNames,states,ignore):
    dirName = "outputFiles"
    fileName = "PCF_App_Details.xlsx"
    sheetName = "Application_Details"
    os.chdir(dirName)
    wb1 = xl.load_workbook(fileName)
    ws = wb1.worksheets[0]
    # print("Available Cell is :", (currRow, currCol))


    # Writing Experience in the File as Header
    if(ws.cell(row=1, column=1).value == None):
        ws.cell(row=1, column=1).value = exp
        currRow = 2
    else:
        currRow = 2

    currCol = getNextAvailableCell()
    currCol += 1

    if currRow == 0 and currCol == 0:
        print("File is Empty")

    # Writing Sub Experience in the File
    if (ignore == False):
        ws.cell(row=currRow, column=currCol).value = subExp

    currRow += 1

    # Writing Org Name in the File
    ws.cell(row=currRow, column=currCol).value = orgName
    currRow += 1

    mName = []
    for managerName in managerNames:
        mName.append(managerName)
    ws.cell(row=currRow, column=currCol).value = mName.__str__().replace("['","").replace("']","").replace("'","")
    currRow += 1

    stateCol = currCol + 1
    stateRow = currRow
    for appName in appNames:
        ws.cell(row=currRow, column=currCol).value = appName
        currRow += 1


    currRow = stateRow
    currCol = stateCol
    for state in states:
        ws.cell(row=currRow, column=currCol).value = state
        currRow += 1
    print("last Row :",currRow)
    print("last Column :", currCol)

    #  Merging Cell to display Header
    #ws.merge_cells(None,1,1,1,currCol)


    wb1.save(fileName)


def getOrgAppDetails():
    dirName = "outputFiles"
    fileName = "//PCF_App_Details.xlsx"
    sheetName = "Application_Details"
    prevCounter = -1
    ignore = False
    cwd = os.getcwd()
    os.chdir(dirName)
    wb = Workbook()  # Workbook is created
    sheetExist = os.path.exists(os.getcwd() + fileName)
    os.chdir(cwd)

    # Call function to get parameter details
    exp, subExp, orgArray = getParameterDetails()
    # Loading Template to call GraphQl
    os.chdir("templates")
    if os.path.exists(os.getcwd() + '//OrgTemplate'):
        TemplateFile = os.getcwd() + '//OrgTemplate'
    templateFileData = open(TemplateFile)
    src = Template(templateFileData.read())

    # Just for fun, display purpose
    star = 0
    for i in orgArray:
        for j in i:
            star = star + 1

    for counter ,orgNames in enumerate(orgArray):
        for orgName in orgNames:
            substituteQuery = {'t1_orgName': '"{0}"'.format(orgName)}
            finalQuery = src.substitute(substituteQuery)
            os.chdir(cwd)

            # Executing GraphQl and fetching data from PCF
            rawData = runQuery(finalQuery)

            # If file exist then it will be overwritten
            if (sheetExist == False):
                createNewFile(dirName, fileName, sheetName)  # New Excel file will be created on given Directory with File and Sheet name
                sheetExist = True
                os.chdir(cwd)

            # Fetching Manager names from extracted pcf data
            managerNames = extract_general_values(rawData, 'managerName')
            print()

            # Fetching Application names from extracted pcf data
            appNames = extract_general_values(rawData, 'appName')

            # Fetching state of the respective application from extracted pcf data
            states = extract_general_values(rawData, 'state')

            prevCounter += 1
            if (counter != prevCounter):
                prevCounter -= 1
                ignore = True
            print("R", prevCounter, "C", counter)
            # Writing file with Appropriate details
            writingFile(exp,orgName, subExp[counter], managerNames, appNames,states,ignore)
            os.chdir(cwd)
            ignore = False

            print("* " * star)
            star -= 1





    '''
    with open("src/inputFile.txt") as fp:
        orgNames = fp.readlines()
        counter = len(orgNames)
        orgNames = list(map(lambda x: x.strip(), orgNames))
        os.chdir("templates")
        if os.path.exists(os.getcwd() + '//OrgTemplate'):
            TemplateFile = os.getcwd() + '//OrgTemplate'
        templateFileData = open(TemplateFile)
        src = Template(templateFileData.read())
        for orgName in orgNames:
            substituteQuery = {'t1_orgName': '"{0}"'.format(orgName)}
            finalQuery = src.substitute(substituteQuery)
            os.chdir(cwd)
            rawData = runQuery(finalQuery)

            if (sheetExist == False):
                createNewFile(dirName, fileName, sheetName)  # New Excel file will be created on given Directory with File and Sheet name
                sheetExist = True
                os.chdir(cwd)

            subExp = getSubExperience(orgName)
            os.chdir(cwd)
            managerNames = extract_values(rawData, 'managerName')
            appNames = extract_values(rawData, 'appName')
            states = extract_values(rawData, 'state')
            writingFile(orgName, subExp, managerNames, appNames,states)
            os.chdir(cwd)
        '''
